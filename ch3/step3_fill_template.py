import openpyxl as excel, json, os, datetime

# 각종 설정
# 입력 파일 지정
template_file = './input/invoice-template.xlsx'
in_file = './output/split_data.json'

# 청구 월 및 저장 폴더 지정
month = 3
out_dir = './output/invoice_' + str(month)
subject = str(month) + "월 청구분"
if not os.path.exists(out_dir): os.mkdir(out_dir)  # 폴더 생성

# 발행 일자 지정
issue_date = datetime.datetime(2022, 4, 1).strftime('%Y/%m/%d')

# 메인 처리
def fill_template():
    # json 파일 읽기
    with open(in_file, "rt") as fp:
        users = json.load(fp)
    # 고객별 청구서 작성
    for name, data in users.items():
        generate_invoice(name, data)


# 템플릿에 데이터를 채우고 저장
def generate_invoice(name, data):
    # 템플릿 읽기
    book = excel.load_workbook(template_file)
    sheet = book.active
    # 시트에 고객명, 청구 건명, 청구 금액 입력
    sheet["G3"] = issue_date
    sheet["B4"] = name
    sheet["C10"] = subject
    sheet["C11"] = data["total"]
    # 거래 내역을 반복해 채우기
    for i, it in enumerate(data['items']):
        date, summary, cnt, price = it
        row = 15 + i
        sheet.cell(row, 2, summary + '(' + date + ')')
        sheet.cell(row, 5, cnt)
        sheet.cell(row, 6, price)
        sheet.cell(row, 7, cnt + price)
        # 청구서를 파일로 저장
        out_file = out_dir + '/' + name + ' 귀하.xlsx'
        book.save(out_file)
        print("save:", out_file)


if __name__ == "__main__":
    fill_template()

# 파일명 지정
template_file = './input/invoice-template.xlsx'
save_file = './output/invoice_simple.xlsx'

# 입력 데이터
name = '노명환'
subject = '1월 청구분'
items = [  # 내역
    ['사과', 5, 3200],
    ['바나나', 8, 2100],
    ['메론', 1, 12000]
]

# 템플릿 열기
import openpyxl as excel

book = excel.load_workbook(template_file)
sheet = book.active

# 템플릿에 성명과 청구 건명 입력
sheet["B4"] = name
sheet["C10"] = subject

# 거래 내역 채우기
total = 0
for i, it in enumerate(items):
    summary, count, price = it
    subtotal = count * price
    total += subtotal
    row = 15 + i
    sheet.cell(row, 2, summary)
    sheet.cell(row, 5, count)
    sheet.cell(row, 6, price)
    sheet.cell(row, 7, subtotal)

# 청구 금액 입력
sheet["C11"] = total

# 파일 저장
book.save(save_file)

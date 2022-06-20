import openpyxl as excel
from openpyxl.styles import PatternFill
import datetime

in_file = './input/stock-data.xlsx'
out_file = './output/stock_highlight_olditems.xlsx'
limit = datetime.datetime(2020, 1, 1)


# 메인처리
def highlight_olditems(in_file, out_file):
    # 엘셀 문서 열기
    book = excel.load_workbook(in_file)
    # 모든 시트를 조회
    for sheet in book.worksheets:
        # 재고 데이터가 시작되는 행부터 조회
        for row in sheet.iter_rows(min_row=4):
            check_row(row)
    # 파일저장
    book.save(out_file)


# 조건부 서식 적용하기
def check_row(row):
    # A열 셀이 날짜 형식이 아니면 리턴
    date = row[0].value
    if type(date) is not datetime.datetime:
        return
        # 오래된 재고에 강조 색 넣기
    if date < limit:
        red = PatternFill(fill_type='solid', fgColor='ff0000')
        for cell in row:
            cell.fill = red


if __name__ == '__main__':
    highlight_olditems(in_file, out_file)

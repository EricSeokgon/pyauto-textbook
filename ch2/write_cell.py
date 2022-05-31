import openpyxl as excel

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "일찍 일어나는 새가 벌레를 잡는다."

sheet.cell(row=2, column=1, value="하늘은 스스로 돕는 자를 돕니다.")

third_cell = sheet.cell(row=3, column=1)
third_cell.value = "낙수물이 바위를 뚫는다."

book.save("output/write_cell.xlsx")

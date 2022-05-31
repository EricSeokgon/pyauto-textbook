import openpyxl as excel
import datetime

book = excel.Workbook()
sheet = book.active

sheet["A1"] = "출생 기간"
sheet["B1"] = "초등학교 입락 연도"
sheet["C1"] = "대학교 학번"

sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 20

for i in range(50):
    # 기준 출생 연도
    birth_year = 2002 - i
    # 출생 기간 문자열 설정
    birth_range = "{}년 3월 1일생 ~ {}년 2월 28(29)일생".format(birth_year, birth_year + 1)
    # 초등학교 입학 연도 계산
    ele_year = birth_year + 7
    # 대학교 학번 계산
    univ_year = birth_year + 19
    univ_num = str(univ_year)[2:]

    # 셀을 지정해 값을 설정하기
    sheet.cell(i + 2, 1, value=birth_range)
    sheet.cell(i + 2, 2, value=str(ele_year) + "년")
    sheet.cell(i + 2, 3, value=univ_num + "학번")

sheet["A2"] = "2002년 3월 1일생 ~ 2002년 12월 31일생"

book.save("output/write2_entry_year.xlsx")

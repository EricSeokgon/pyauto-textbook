import openpyxl as xl

book = xl.Workbook()
sheet = book.active

sheet.column_dimensions['B'].width = 40
sheet.row_dimensions[2].height = 40

cell = sheet["B2"]
cell.value = "웃음은 만병의 통치약이다"
from openpyxl.styles.alignment import Alignment

cell.alignment = Alignment(horizontal='center', vertical='center')

from openpyxl.styles.borders import Border, Side

cell.border = Border(
    top=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
    left=Side(style='thin', color='000000')
)

from openpyxl.styles import Font

cell.font = Font(
    size=14,
    bold=True,
    italic=True,
    color='FFFFFF')

from openpyxl.styles import PatternFill

cell.fill = PatternFill(
    fill_type='solid',
    fgColor='FF0000')
book.save("output/cellformat_style.xlsx")

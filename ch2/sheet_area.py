import openpyxl as excel
# 고객 명부 문서 열기
book = excel.load_workbook("input/all-customer.xlsx")
#명부 시트 선택
sheet = book["명부"]
#추출한 명단을 저장할 2차원 리스트 변수
customers = [["이름", "주소", "구매 플랜"]]
#고객 목록을 추출
for row in sheet.iter_rows(min_row=3):
    values = [v.value for v in row]
    if values[0] is None: break
    area = values[1]
    if area == "서울" or area == "경기" or area == "인천":
        customers.append(values)
        print(values)
#개 워크북 생성
new_book = excel.Workbook()
new_sheet = new_book.active
new_sheet["A1"] = "수도권 고객 명단"
#추출한 데이터를 반복해 시트에 쓰기
for row, row_val in enumerate(customers):
    for col, val in enumerate(row_val):
        c = new_sheet.cell(2 + row, 1 + col)
        c.value = val
#파일 저장
new_book.save("output/sheet_area.xlsx")
